"""XLSX parser (openpyxl read-only streaming): each sheet a table block."""

import io

from openpyxl import load_workbook

from src.knowledge.ingest_errors import ParserError
from src.knowledge.parsing.base import ContentBlock, ParsedDocument


class XlsxParser:
    def parse(self, data: bytes, *, max_pages: int) -> ParsedDocument:
        """Stream rows in read-only mode so memory is bounded by the row width,
        not the sheet size. Time: O(cells). Space: O(one row)."""
        try:
            workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        except Exception as exc:
            raise ParserError("file is not a readable XLSX") from exc
        try:
            blocks: list[ContentBlock] = []
            for worksheet in workbook.worksheets:
                rows = [
                    " | ".join(str(cell) for cell in row if cell is not None)
                    for row in worksheet.iter_rows(values_only=True)
                ]
                non_empty = [row for row in rows if row]
                if non_empty:
                    blocks.append(
                        ContentBlock(
                            text="\n".join(non_empty),
                            ref={"sheet": worksheet.title},
                            is_table=True,
                        )
                    )
            page_count = len(workbook.sheetnames) or 1
        finally:
            workbook.close()
        return ParsedDocument(
            blocks=tuple(blocks), page_count=page_count, table_ratio=1.0 if blocks else 0.0
        )
